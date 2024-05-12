import pandas as pd
import seaborn as sns
import numpy as np
import matplotlib.pyplot as plt
import csv
import os
import copy
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

# Function to retrieve Zoom meeting info from CERN INDICO website
def retrieve_zoom_info(username, password, page_url):
    session = requests.Session()

    # Login
    login_data = {"username": username, "password": password}
    session.post(login_url, data=login_data)

    # Request the page after login
    response = session.get(page_url)

    # Parse HTML
    soup = BeautifulSoup(response.content, "html.parser")

    # Find all elements containing Zoom links
    zoom_meetings = soup.find_all("div", class_="event-service-row")

    # Prepare data
    zoom_info = {}
    zoom_links = []
    zoom_meetingid = []
    zoom_passcode = []
    zoom_event_name = []
    for meeting in zoom_meetings:
        event_name = meeting.find("span", class_="event-service-title").text.strip()
        meeting_id = meeting.find("dt", text="Zoom Meeting ID").find_next_sibling("dd").text.strip()
        passcode = meeting.find("dt", text="Passcode").find_next_sibling("dd").text.strip()
        zoom_url = meeting.find("input", type="text")["value"]
        zoom_info[event_name] = {"Meeting ID": meeting_id, "Passcode": passcode, "Zoom URL": zoom_url}
        zoom_event_name.append(event_name)
        zoom_links.append(zoom_url)
        zoom_meetingid.append(meeting_id)
        zoom_passcode.append(passcode)

    return zoom_event_name, zoom_meetingid, zoom_links, zoom_passcode

# Credentials
username = ""
password = ""

# URLs
login_url = "https://auth.cern.ch/auth/realms/cern/login-actions/authenticate?execution=133f73d5-6454-4197-b529-b109a5d9432c&client_id=indico-cern&tab_id=tgmlNdWhv-o"
#2023
#page_url = "https://indico.cern.ch/event/1391268/videoconference/"
#2024
page_url = "https://indico.cern.ch/event/1391268/videoconference/"

# Retrieve Zoom meeting info and links
zoom_event_name, zoom_meetingid, zoom_links, zoom_passcode = retrieve_zoom_info(username, password, page_url)

#If true, debug printouts are enabled
debug = False

#list of recruiters and time of their session
#2023
"""
recruiters = ["Manjarres","Roloff","Kortner","Williams","Balli","Corpe","Roloff","Williams","Evans","Manjarres","Kortner","Mijovic","Sofonov","Stupak","Haas","Mijovic","Sofonov","Haas","Whiteson","Evans","Corpe","Balli","Orimoto","Stupak","Orimoto"]
times = ["09:00-10:00","09:00-10:00","10:00-11:00","10:00-11:00","11:00-12:00","11:00-12:00","12:00-13:00","12:00-13:00","14:00-15:00","14:00-15:00","14:00-15:00","15:00-15:55","15:00-16:00","15:00-16:00","15:30-16:30","16:00-17:00","16:05-17:05","16:35-17:35","17:00-18:00","17:00-18:00","17:00-18:00","18:00-19:00","18:00-18:55","19:00-20:00","19:00-20:00"]
"""
#2024
recruiters = ["Boonekamp","Shu Li","Spousta","Kenzie","Mengqing Wu","D'Eramo","Sculac","Boonekamp","Munhoz","Muskinja","Spousta","Gouskos","Arguin","D'Eramo","Shu Li","Simon","Munhoz","Muskinja","Porteboeuf"]
times = ["09:00-09:55","09:00-09:55","10:00-10:55","11:00-11:55","11:00-11:55","12:00-12:55","12:00-12:55","14:00-14:55","14:00-14:55","15:00-15:55","15:00-15:55","15:00-15:55","16:00-16:55","16:00-16:55","16:00-16:55","16:00-16:55","17:00-17:55","17:00-17:55","17:00-17:55"]


#list to keep track of total participants to a given session, initialized to 0's according to how many sessions are foreseen
totParticipants = []
for i in range(len(recruiters)):
    totParticipants.append(0)

#open the list of people as pandas df
#jobMatching = pd.read_csv("registrations_jobMatching_2023.csv")
jobMatching = pd.read_csv("registrations2024.csv")

#Filter the df if there are empty "cells" (for example if a cell is empty it's read as NaN and it messes up the code later on)
jobMatchingFiltered = jobMatching.dropna()

jobOfferSessionParticipation = []

#Participant names and sessions as a list, extracted from the df 
participants = jobMatchingFiltered.loc[:,"Name"].tolist()
sessions = jobMatchingFiltered.loc[:,"Job Offer Sessions"].tolist()

#Some participants had not indicated their sessions and it said this phrase: To be updated by Oct 17th once recruiter list is finalized, please check back then to update your registration. :)
#So I remove it and I also remove the name of the participants associated with it
for i,j in enumerate(participants):
    if sessions[i] == "To be updated by Oct 17th once recruiter list is finalized, please check back then to update your registration. :)":
        participants.remove(participants[i])
        sessions.remove(sessions[i])

#There is also the case that for some people one of the session is indicated as To be updated [...]
#I also remove these ones
for i in sessions:
    if debug:
        print("analyzing")
    participantSessions = i.split(";")
    for x in participantSessions:
        if x == ' To be updated by Oct 17th once recruiter list is finalized, please check back then to update your registration. :)' or x == ' A session I want to attend is full.':
            participantSessions.remove(x)
                
    jobOfferSessionParticipation.append(participantSessions)

#merge names of recruiters and time of the session to use a title of the columns
mergedTimeRecruiters = [m + " " + str(n) for m,n in zip(recruiters,times)]
mergedTimeRecruiters.insert(0,"Name")
mergedTimeRecruiters.insert(1,"NOTES")
mergedTimeRecruiters.append("Total sessions")

#list of 0s and 1s for all sessions, one list per participant (if == 1 student is participating to the session, if 0 no)
participating = []

if debug:
    print (len(jobOfferSessionParticipation),len(participants))

#Manipulate zoom links and passwords according to the specific session
ordered_zoom_event_id = []
ordered_zoom_password = []
ordered_zoom_link = []

for r,recruiter in enumerate(recruiters):
    for e,event in enumerate(zoom_event_name):
        if event.find(recruiter) != -1:
            ordered_zoom_event_id.append(zoom_meetingid[e])
            ordered_zoom_password.append(zoom_passcode[e])
            ordered_zoom_link.append(zoom_links[e])
            break

print(ordered_zoom_event_id)

#remove .csv output file if it exists already
if os.path.exists('participantView2024.csv'):
    print("Removing old file")
    os.remove('participantView2024.csv')
else:
    print("File does not exist, moving on")

#Here we write the actual output file
with open('participantView2024.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    
    #This writes the first line (surname of job recruiter and time slot)
    writer.writerow(mergedTimeRecruiters)

    #This checkes for each job seeker if they have applied for a specific session
    for i,studentSession in enumerate(sessions):
        for j,recruiter in enumerate(recruiters):              
            if studentSession.find(recruiters[j]) != -1 and studentSession.find(times[j]) != -1:
                #if yes -> put a 1 in the participating list  
                participating.append(1)
                #increase by 1 the number of participants to the session
                totParticipants[j] += 1
            else:
                #if no append a 0
                participating.append(0)

        totSessionPerParticipant = 0
        totSessionPerParticipant = sum(participating)

        #Once the loop is over add the name of the student as the frist element of the participating list
        participating.insert(0,participants[i])
        participating.insert(1,"")
        participating.append(totSessionPerParticipant)
        #write to the .csv file
        writer.writerow(participating)
        #clear list for next iteration       
        participating.clear()

    #Add name of tot participants list
    totParticipants.insert(0,"Tot participants to session")
    totParticipants.insert(1,"")
    #write total number of participants to each session
    writer.writerow(totParticipants)

    #Add name of the zoom sessions list
    ordered_zoom_event_id.insert(0,"Zoom Meeting ID")
    ordered_zoom_event_id.insert(1,"")
    #write zoom links for each session
    writer.writerow(ordered_zoom_event_id)
    
    #Add name of the zoom sessions list
    ordered_zoom_password.insert(0,"Zoom Passcode")
    ordered_zoom_password.insert(1,"")
    #write zoom links for each session
    writer.writerow(ordered_zoom_password)
    
    #Add name of the zoom sessions list
    ordered_zoom_link.insert(0,"Zoom Links")
    ordered_zoom_link.insert(1,"")
    #write zoom links for each session
    writer.writerow(ordered_zoom_link)

    moderator1 = ["Moderator 1"]
    writer.writerow(moderator1)

    moderator2 = ["Moderator 2 and shadows"]
    writer.writerow(moderator2)

    notes = ["Comments"]
    writer.writerow(notes)

    legend = ["Legend","Attended","Did not attend"]
    writer.writerow(legend)

partView2024 = pd.read_csv("participantView2024.csv")

#remove .xlsx output file if it exists already
if os.path.exists('participantView2024.xlsx'):
    print("Removing old file")
    os.remove('participantView2024.xlsx')
else:
    print("File does not exist, moving on")

partView2024.to_excel('participantView2024.xlsx', sheet_name='sheet1', index=False)

"""
Fancy-up the file
"""
#Open excel version of the file
wb = openpyxl.load_workbook("participantView2024.xlsx")
ws = wb['sheet1']

#Background color of cells
clr_background_first_line = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type="solid") #yellow blue for recruiters and job seekers
clr_background_names = PatternFill(start_color='0099CCFF', end_color='0099CCFF', fill_type="solid") #light blue for recruiters and job seekers
clr_background_white = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid") #white
clr_background_yellow_light = PatternFill(start_color='00FFFF99', end_color='00FFFF99', fill_type="solid") #light yellow
clr_background_red = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type="solid") #red
clr_background_green = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type="solid") #green
clr_background_light_gray = PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type="solid") #light grey


#cell border style
border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

#lock first column and row
ws.freeze_panes = 'B2'

#Go through all the cells and apply style changes
for row in ws.iter_rows():
    for cell in row:

        cell.alignment = Alignment(horizontal='center', vertical='center') #Center content of all cells
        cell.border = border #thin cell border

        if cell.column == 1: #change content of first column (name of participants)
            cell.font = Font(b = True) #bold 
            cell.fill = clr_background_names #bckg color light blue 

        if cell.row == 1:
            cell.fill = clr_background_first_line #color the cells in the first row

        #zoom links as hyperlinks to take up less space
        #try/except is used because if the cell is empty an error is thrown
        try:
            if cell.value.find("https://cern.zoom.us") != -1: #if the cell contains the link -> it means it's a zoom link
                cell.hyperlink = cell.value
                cell.value = 'link'
                cell.style = "Hyperlink"
                #reformat after changing text
                cell.alignment = Alignment(horizontal='center', vertical='center') #Center content of all cells
                cell.border = border #thick cell border

        except:
            pass

        if cell.value == "Tot participants to session":
            cell.font = Font(b = True)  #make cell content bold
            cell.fill = clr_background_green #green background color
        
        elif cell.value == "Zoom Meeting ID" or cell.value == "Zoom Passcode" or cell.value == "Zoom Links":
            cell.font = Font(b = True) #make cell content bold
            cell.fill = clr_background_light_gray #light gray background color
        
        elif cell.value == "Notes":
            cell.fill = clr_background_yellow_light #light yellow background color

        elif cell.value == "Legend":
            cell.fill = clr_background_white #white background color

        elif cell.value == "Attended":
            cell.fill = clr_background_green #green background color

        elif cell.value == "Did not attend":
            cell.fill = clr_background_red #red background color

        elif cell.value == "Comments": #enlarge cell for comments
            nameRow = cell.row
            ws.row_dimensions[nameRow].height = 250

for col in ws.iter_cols():
    nameColumn = get_column_letter(col[0].column)
    new_col_length = max(len(str(cell.value)) for cell in col)
    ws.column_dimensions[nameColumn].width = 30 # Enlarge all columns

#Save file to apply changes
wb.save(filename="participantView2024.xlsx")

#Remove .csv file
if os.path.exists('participantView2024.csv'):
    print("Removing .csv file")
    os.remove('participantView2024.csv')
else:
    print("File does not exist, please check")
